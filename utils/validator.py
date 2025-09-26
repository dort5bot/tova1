#Excel Doğrulayıcı (utils/validator.py)
"""
"TARİH", "İL" doğrulaması yapar

"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import Dict, Any
from utils.logger import logger

def validate_excel_file(file_path: str) -> Dict[str, Any]:
    """
    Excel dosyasını doğrular
    """
    try:
        wb = load_workbook(filename=file_path, read_only=True)
        ws = wb.active
        
        # Başlık satırını al
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            headers.append(str(cell_value).strip().upper() if cell_value else "")
        
        # Gerekli sütunları kontrol et
        required_columns = {"TARİH", "İL"}
        found_columns = set(headers)
        
        if not required_columns.issubset(found_columns):
            missing = required_columns - found_columns
            return {
                "valid": False,
                "message": f"Dosyada gerekli sütunlar bulunamadı: {', '.join(missing)}"
            }
        
        # Satır sayısını kontrol et (sadece başlık varsa)
        if ws.max_row <= 1:
            return {
                "valid": False,
                "message": "Dosyada işlenecek veri bulunamadı"
            }
        
        return {"valid": True, "headers": headers, "row_count": ws.max_row - 1}
        
    except Exception as e:
        logger.error(f"Doğrulama hatası: {e}")
        return {"valid": False, "message": f"Dosya okunamadı: {str(e)}"}
    finally:
        if 'wb' in locals():
            wb.close()
