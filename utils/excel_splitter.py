# utils/excel_splitter.py (güncellenmiş versiyon)
"""
Excel dosyasını gruplara ayıran ana fonksiyon

"""
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from typing import Dict, List, Tuple, Any
import tempfile
import os

from utils.group_manager import group_manager
from utils.file_namer import generate_output_filename
from utils.logger import logger
from config import config

class ExcelSplitter:
    def __init__(self):
        self.workbooks = {}  # group_id -> Workbook
        self.sheets = {}     # group_id -> Worksheet
        self.row_counts = {} # group_id -> satır sayısı
        self.headers = []    # başlık satırı
        self.city_mapping_stats = {}  # Şehir eşleştirme istatistikleri
    
    def initialize_workbook(self, group_id: str):
        """Yeni bir workbook ve worksheet oluşturur"""
        if group_id not in self.workbooks:
            wb = Workbook()
            ws = wb.active
            ws.title = "Veriler"
            
            # Başlık satırını yaz
            for col_idx, header in enumerate(self.headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            # Sütun genişliklerini ayarla (25 birim)
            self.adjust_column_widths(ws)
            
            self.workbooks[group_id] = wb
            self.sheets[group_id] = ws
            self.row_counts[group_id] = 1  # Başlık satırı
            self.city_mapping_stats[group_id] = 0
    
    def adjust_column_widths(self, worksheet, width: int = 25):
        """
        Tüm sütunların genişliğini ayarlar
        Args:
            worksheet: Çalışma sayfası nesnesi
            width: Sütun genişliği (varsayılan: 25)
        """
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            column_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[column_letter].width = min(width, max(length + 2, 10))
    
    def process_excel_file(self, input_path: str, headers: List[str]) -> Dict[str, Any]:
        """Excel dosyasını gruplara ayırır (optimize edilmiş)"""
        wb = None  # Workbook'u burada tanımla
        try:
            self.headers = headers
            self.city_mapping_stats = {}
            
            # Read-only modunda aç (context manager OLMADAN)
            wb = load_workbook(filename=input_path, read_only=True)
            ws = wb.active
                
            total_rows = ws.max_row - 1  # Başlık hariç
            logger.info(f"İşlenecek toplam satır: {total_rows}")
            
            # Tüm satırları iterate et
            processed_rows = 0
            unmatched_cities = set()
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Boş satırları atla
                    continue
                
                # İl bilgisini al (B sütunu - index 1)
                city = row[1] if len(row) > 1 else None
                
                # Grubu belirle (birden fazla grup olabilir)
                group_ids = group_manager.get_groups_for_city(city)
                
                if "Grup_0" in group_ids and len(group_ids) == 1:
                    unmatched_cities.add(str(city))
                
                # Her grup için satırı ekle
                for group_id in group_ids:
                    self.initialize_workbook(group_id)
                    
                    ws_dest = self.sheets[group_id]
                    current_row = self.row_counts[group_id] + 1
                    
                    for col_idx, value in enumerate(row, 1):
                        ws_dest.cell(row=current_row, column=col_idx, value=value)
                    
                    self.row_counts[group_id] = current_row
                    self.city_mapping_stats[group_id] += 1
                
                processed_rows += 1
                
                # İlerleme logu (her 1000 satırda bir)
                if processed_rows % 1000 == 0:
                    logger.info(f"{processed_rows}/{total_rows} satır işlendi")
            
            logger.info(f"İşlem tamamlandı: {processed_rows} satır")
            
            # Eşleşmeyen şehirleri logla
            if unmatched_cities:
                logger.warning(f"Eşleşmeyen şehirler: {list(unmatched_cities)[:10]}{'...' if len(unmatched_cities) > 10 else ''}")
            
            # Dosyaları kaydetmeden önce sütun genişliklerini güncelle
            for group_id, wb in self.workbooks.items():
                if self.row_counts[group_id] > 1:  # Sadece başlık değilse
                    self.adjust_column_widths(self.sheets[group_id])
            
            # Dosyaları kaydet
            output_files = {}
            for group_id, wb in self.workbooks.items():
                if self.row_counts[group_id] > 1:  # Sadece başlık değilse
                    group_info = group_manager.get_group_info(group_id)
                    filename = generate_output_filename(group_info)
                    filepath = config.OUTPUT_DIR / filename
                    
                    # Dizin yoksa oluştur
                    filepath.parent.mkdir(parents=True, exist_ok=True)
                    
                    wb.save(filepath)
                    output_files[group_id] = {
                        "path": filepath,
                        "row_count": self.row_counts[group_id] - 1,  # Başlık hariç
                        "filename": filename,
                        "matched_cities": self.city_mapping_stats.get(group_id, 0)
                    }
            
            matched_rows = sum(count - 1 for count in self.row_counts.values() if count > 1)
            
            return {
                "success": True,
                "output_files": output_files,
                "total_rows": processed_rows,
                "matched_rows": matched_rows,
                "unmatched_cities": list(unmatched_cities),
                "stats": self.city_mapping_stats
            }
            
        except Exception as e:
            logger.error(f"Excel ayırma hatası: {e}", exc_info=True)
            return {"success": False, "error": str(e)}
        finally:
            # Workbook'u kapat (eğer açıksa)
            if wb is not None:
                try:
                    wb.close()
                except:
                    pass
            
            # Belleği temizle
            self.close_all_workbooks()
    
    def close_all_workbooks(self):
        """Tüm workbook'ları kapatır"""
        for wb in self.workbooks.values():
            try:
                wb.close()
            except:
                pass
        self.workbooks.clear()
        self.sheets.clear()
        self.row_counts.clear()
        self.city_mapping_stats.clear()

def split_excel_by_groups(input_path: str, headers: List[str]) -> Dict[str, Any]:
    """Excel dosyasını gruplara ayıran ana fonksiyon"""
    splitter = ExcelSplitter()
    return splitter.process_excel_file(input_path, headers)