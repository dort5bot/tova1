#Excel Temizleyici (utils/excel_cleaner.py)
# sütun genişliği ayarı eklendi
#Sütun genişliği otomatik olarak içeriğe göre ayarlanır, minimum 10, maksimum 25 birim


from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import Dict, List, Tuple, Any
from utils.logger import logger
import tempfile
import os

def clean_excel_headers(input_path: str) -> Dict[str, Any]:
    """
    Excel dosyasının başlıklarını temizler ve düzenler
    """
    try:
        wb = load_workbook(filename=input_path)
        ws = wb.active
        
        # Başlık satırını bul (boş satırları atlayarak)
        header_row = 1
        for row in range(1, 6):  # İlk 5 satırı kontrol et
            if any(ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)):
                header_row = row
                break
        
        # Başlıkları al ve temizle
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            clean_value = str(cell_value).strip().upper() if cell_value else f"UNKNOWN_{col}"
            headers.append(clean_value)
        
        # TARİH ve İL sütunlarının indekslerini bul
        date_idx = None
        city_idx = None
        
        for idx, header in enumerate(headers, 1):
            if "TARİH" in header:
                date_idx = idx
            elif "İL" in header and city_idx is None:
                city_idx = idx
        
        if date_idx is None or city_idx is None:
            raise ValueError("TARİH veya İL sütunu bulunamadı")
        
        # Sütunları yeniden düzenle: TARİH -> A, İL -> B, diğerleri C'den itibaren
        new_headers = ["TARİH", "İL"]
        other_headers = []
        
        for idx, header in enumerate(headers, 1):
            if idx not in [date_idx, city_idx] and header not in new_headers:
                other_headers.append(header)
        
        new_headers.extend(other_headers)
        
        # Yeni bir çalışma sayfası oluştur
        # Yeni bir çalışma sayfası oluştur
        from openpyxl import Workbook
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = "Düzenlenmiş Veri"
        
        # Yeni başlıkları yaz
        for col_idx, header in enumerate(new_headers, 1):
            new_ws.cell(row=1, column=col_idx, value=header)
        
        # Verileri taşı
        new_row_idx = 2
        for row in range(header_row + 1, ws.max_row + 1):
            # TARİH sütunundan veriyi al
            date_value = ws.cell(row=row, column=date_idx).value
            # İL sütunundan veriyi al
            city_value = ws.cell(row=row, column=city_idx).value
            
            # Yeni satıra TARİH ve İL değerlerini yaz
            new_ws.cell(row=new_row_idx, column=1, value=date_value)
            new_ws.cell(row=new_row_idx, column=2, value=city_value)
            
            # Diğer sütunları kopyala
            new_col_idx = 3
            for col in range(1, ws.max_column + 1):
                if col not in [date_idx, city_idx]:
                    value = ws.cell(row=row, column=col).value
                    new_ws.cell(row=new_row_idx, column=new_col_idx, value=value)
                    new_col_idx += 1
            
            new_row_idx += 1
        
        
        # Sütun genişliklerini ayarla
        from openpyxl.utils import get_column_letter
        for column_cells in new_ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            column_letter = get_column_letter(column_cells[0].column)
            new_ws.column_dimensions[column_letter].width = min(25, max(length + 2, 10))
        
        
                
        # Geçici dosyaya kaydet
        # Geçici dosyaya kaydet
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        temp_path = temp_file.name
        new_wb.save(temp_path)
        
        return {
            "success": True,
            "temp_path": temp_path,
            "headers": new_headers,
            "row_count": new_row_idx - 2
        }
        
    except Exception as e:
        logger.error(f"Excel temizleme hatası: {e}")
        return {"success": False, "error": str(e)}
    finally:
        if 'wb' in locals():
            wb.close()
        if 'new_wb' in locals():
            new_wb.close()
