# utils/json_processing.py
# openpyxl ile
import os
import json
import logging
import aiofiles
import asyncio
from openpyxl import load_workbook
from typing import Dict, List, Any

logger = logging.getLogger(__name__)

async def process_excel_to_json(excel_file_path: str) -> str:
    """
    Excel dosyasını işleyerek groups.json dosyası oluşturur.
    
    Args:
        excel_file_path: İşlenecek Excel dosyasının yolu
        
    Returns:
        Oluşturulan JSON dosyasının yolu
    """
    try:
        # Excel dosyasını senkron olarak yükle (openpyxl async desteklemiyor)
        def load_excel_sync():
            return load_workbook(excel_file_path, read_only=True)
        
        # Thread pool'da Excel yükleme
        loop = asyncio.get_event_loop()
        wb = await loop.run_in_executor(None, load_excel_sync)
        
        # "grup" sayfasını kontrol et
        if 'grup' not in wb.sheetnames:
            raise ValueError("Excel dosyasında 'grup' sayfası bulunamadı")
        
        ws = wb['grup']
        
        # Grup verilerini topla
        groups_data = extract_groups_data(ws)
        
        # Çıktı dizinini oluştur
        output_dir = "data/groups"
        os.makedirs(output_dir, exist_ok=True)
        
        # JSON dosya yolu
        json_file_path = os.path.join(output_dir, "groups.json")
        
        # JSON'ı asenkron olarak yaz
        async with aiofiles.open(json_file_path, 'w', encoding='utf-8') as f:
            json_data = json.dumps({"groups": groups_data}, ensure_ascii=False, indent=2)
            await f.write(json_data)
        
        logger.info(f"JSON dosyası başarıyla oluşturuldu: {json_file_path}")
        return json_file_path
        
    except Exception as e:
        logger.error(f"Excel işleme hatası: {str(e)}", exc_info=True)
        raise
    finally:
        # Workbook'u kapat
        if 'wb' in locals():
            wb.close()

def get_column_letter(n: int) -> str:
    """
    Sütun numarasını Excel harfine çevirir.
    Örnek: 1 -> 'A', 27 -> 'AA'
    """
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result

def extract_groups_data(worksheet) -> List[Dict[str, Any]]:
    """
    Worksheet'ten grup verilerini çıkarır.
    
    Args:
        worksheet: Openpyxl worksheet objesi
        
    Returns:
        Grup verileri listesi
    """
    groups = []
    
    # Sütunları D'dan başlayarak tarayın (sütun 4)
    column_index = 4  # D sütunu = 4
    
    while True:
        # Sütun harfini al (AA, AB, ... dahil)
        column_letter = get_column_letter(column_index)
        
        # Grup ID kontrolü (1. satır)
        group_id_cell = f"{column_letter}1"
        group_id = worksheet[group_id_cell].value if worksheet[group_id_cell].value else None
        
        # Boş sütun bulunursa dur
        if not group_id:
            break
        
        # Grup adı (2. satır)
        group_name_cell = f"{column_letter}2"
        group_name = worksheet[group_name_cell].value if worksheet[group_name_cell].value else ""
        
        # E-posta listesi (3. satır)
        email_cell = f"{column_letter}3"
        email_recipients = worksheet[email_cell].value if worksheet[email_cell].value else ""
        
        # Şehirleri topla (4. satırdan itibaren)
        cities = []
        row_index = 4
        
        while True:
            city_cell = f"{column_letter}{row_index}"
            city = worksheet[city_cell].value
            
            # Boş hücre bulunursa şehirleri toplamayı durdur
            if not city:
                break
            
            cities.append(str(city).strip())
            row_index += 1
        
        # E-postaları temizle ve liste olarak ayır
        email_list = []
        if email_recipients:
            email_list = [email.strip() for email in str(email_recipients).split(',') if email.strip()]
        
        # Grup verisini oluştur
        group_data = {
            "group_id": str(group_id).strip(),
            "group_name": str(group_name).strip(),
            "email_recipients": email_list,
            "cities": cities
        }
        
        groups.append(group_data)
        column_index += 1  # Sonraki sütuna geç
    
    return groups
