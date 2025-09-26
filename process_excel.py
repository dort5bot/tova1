# Excel İşleme Görevi (jobs/process_excel.py)

import asyncio
from pathlib import Path
from typing import Dict, Any, List
from openpyxl import load_workbook
import tempfile

from utils.excel_cleaner import clean_excel_headers
from utils.excel_splitter import split_excel_by_groups
from utils.mailer import send_email_with_attachment
from utils.group_manager import group_manager
from utils.logger import logger
import config  # config modülünü import etmeyi unutmayın

async def process_excel_task(input_path: Path, user_id: int) -> Dict[str, Any]:
    """Excel işleme görevini yürütür (geliştirilmiş)"""
    cleaning_result = None
    try:
        logger.info(f"Excel işleme başlatıldı: {input_path.name}, Kullanıcı: {user_id}")

        # 1. Excel dosyasını temizle ve düzenle
        cleaning_result = clean_excel_headers(str(input_path))
        if not cleaning_result["success"]:
            error_msg = f"Excel temizleme hatası: {cleaning_result.get('error', 'Bilinmeyen hata')}"
            logger.error(error_msg)
            return {"success": False, "error": error_msg}
        
        logger.info(f"Excel temizlendi: {cleaning_result['row_count']} satır")

        # 2. Dosyayı gruplara ayır
        splitting_result = split_excel_by_groups(
            cleaning_result["temp_path"],
            cleaning_result["headers"]
        )
        
        if not splitting_result["success"]:
            error_msg = f"Excel ayırma hatası: {splitting_result.get('error', 'Bilinmeyen hata')}"
            logger.error(error_msg)
            return {"success": False, "error": error_msg}
        
        logger.info(f"Excel gruplara ayrıldı: {splitting_result['total_rows']} satır, {len(splitting_result['output_files'])} grup")

        # 3. E-postaları gönder (async olarak)
        email_tasks = []
        output_files = splitting_result["output_files"]
        email_results = []
        
        for group_id, file_info in output_files.items():
            group_info = group_manager.get_group_info(group_id)
            recipients = group_info.get("email_recipients", [])
            
            if recipients and file_info["row_count"] > 0:
                subject = f"{group_info.get('group_name', group_id)} Raporu - {file_info['filename']}"
                body = (
                    f"Merhaba,\n\n"
                    f"{group_info.get('group_name', group_id)} grubu için {file_info['row_count']} satırlık rapor ekte gönderilmiştir.\n\n"
                    f"İyi çalışmalar,\nExcel Bot"
                )
                
                # Her alıcı için ayrı mail gönderimi
                for recipient in recipients:
                    if recipient.strip():  # Boş email adreslerini atla
                        task = send_email_with_attachment(
                            [recipient.strip()], subject, body, file_info["path"]
                        )
                        email_tasks.append((task, group_id, recipient, file_info["path"].name))
        
        # Tüm mail görevlerini paralel çalıştır
        if email_tasks:
            logger.info(f"{len(email_tasks)} mail görevi başlatılıyor...")
            
            # Görevleri topla ve çalıştır
            tasks = [task[0] for task in email_tasks]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            # Sonuçları işle
            for i, result in enumerate(results):
                task_info = email_tasks[i]
                group_id, recipient, filename = task_info[1], task_info[2], task_info[3]
                
                if isinstance(result, Exception):
                    logger.error(f"Mail gönderim hatası - Grup: {group_id}, Alıcı: {recipient}, Dosya: {filename}, Hata: {result}")
                    email_results.append({
                        "success": False,
                        "group_id": group_id,
                        "recipient": recipient,
                        "error": str(result)
                    })
                else:
                    logger.info(f"Mail gönderildi - Grup: {group_id}, Alıcı: {recipient}, Dosya: {filename}")
                    email_results.append({
                        "success": True,
                        "group_id": group_id,
                        "recipient": recipient
                    })
            
            # Grup bazında email durumunu güncelle
            successful_emails = sum(1 for res in email_results if res["success"])
            logger.info(f"Mail gönderim sonucu: {successful_emails} başarılı, {len(email_results) - successful_emails} başarısız")
        
        # 4. Geçici dosyaları temizle
        try:
            if cleaning_result and "temp_path" in cleaning_result:
                temp_path = Path(cleaning_result["temp_path"])
                if temp_path.exists():
                    temp_path.unlink()
                    logger.info(f"Geçici dosya silindi: {temp_path.name}")
        except Exception as e:
            logger.warning(f"Geçici dosya silinemedi: {e}")
        
        return {
            "success": True,
            "output_files": output_files,
            "total_rows": splitting_result["total_rows"],
            "matched_rows": splitting_result["matched_rows"],
            "email_results": email_results,
            "user_id": user_id
        }
        
    except Exception as e:
        logger.error(f"İşlem görevi hatası: {e}", exc_info=True)
        
        # Hata durumunda geçici dosyaları temizle
        try:
            if cleaning_result and "temp_path" in cleaning_result:
                temp_path = Path(cleaning_result["temp_path"])
                if temp_path.exists():
                    temp_path.unlink()
        except:
            pass
            
        return {"success": False, "error": str(e)}
        



#kişisel mail fonksiyonu
#
async def process_excel_task_for_personal_email(input_path: Path, user_id: int) -> Dict[str, Any]:
    """Sadece kişisel maile gönderim için Excel işleme görevi"""
    cleaning_result = None
    try:
        logger.info(f"Kişisel mail için Excel işleme başlatıldı: {input_path.name}, Kullanıcı: {user_id}")

        # 1. Excel dosyasını temizle ve düzenle (aynı işlem)
        cleaning_result = clean_excel_headers(str(input_path))
        if not cleaning_result["success"]:
            error_msg = f"Excel temizleme hatası: {cleaning_result.get('error', 'Bilinmeyen hata')}"
            logger.error(error_msg)
            return {"success": False, "error": error_msg}
        
        logger.info(f"Excel temizlendi: {cleaning_result['row_count']} satır")

        # 2. Tüm verileri tek bir dosyada tut (gruplara ayırma YOK)
        # Temizlenmiş dosyayı yükle
        wb = load_workbook(cleaning_result["temp_path"])
        ws = wb.active
        
        # Sütun genişliklerini ayarla
        from openpyxl.utils import get_column_letter
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            column_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[column_letter].width = min(25, max(length + 2, 10))
        
        # Geçici çıktı dosyası oluştur
        temp_output = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        temp_output_path = temp_output.name
        wb.save(temp_output_path)
        wb.close()
        
        # 3. Sadece kişisel maile gönder
        email_success = False
        if config.PERSONAL_EMAIL:
            subject = f"📊 Excel Raporu - {input_path.name}"
            body = (
                f"Merhaba,\n\n"
                f"{cleaning_result['row_count']} satırlık Excel raporu ekte gönderilmiştir.\n\n"
                f"İyi çalışmalar,\nExcel Bot"
            )
            
            email_success = await send_email_with_attachment(
                [config.PERSONAL_EMAIL], subject, body, Path(temp_output_path)
            )
        
        # 4. Geçici dosyaları temizle
        try:
            if cleaning_result and "temp_path" in cleaning_result:
                temp_path = Path(cleaning_result["temp_path"])
                if temp_path.exists():
                    temp_path.unlink()
            if Path(temp_output_path).exists():
                Path(temp_output_path).unlink()
        except Exception as e:
            logger.warning(f"Geçici dosya silinemedi: {e}")
        
        return {
            "success": email_success,
            "total_rows": cleaning_result["row_count"],
            "email_sent_to": config.PERSONAL_EMAIL if email_success else None,
            "user_id": user_id
        }
        
    except Exception as e:
        logger.error(f"Kişisel mail işleme hatası: {e}", exc_info=True)
        
        # Hata durumunda geçici dosyaları temizle
        try:
            if cleaning_result and "temp_path" in cleaning_result:
                temp_path = Path(cleaning_result["temp_path"])
                if temp_path.exists():
                    temp_path.unlink()
            if 'temp_output_path' in locals() and Path(temp_output_path).exists():
                Path(temp_output_path).unlink()
        except:
            pass
            
        return {"success": False, "error": str(e)}